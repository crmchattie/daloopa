from datetime import datetime
from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
import json
from pymongo import MongoClient
import os

app = FastAPI()

# Predefined list of sections
sections = [
    "Document", 
    "KPIs", 
    "Geography Breakdown", 
    "Income Statement", 
    "Balance Sheet", 
    "Cash Flow Statement", 
    "Adjusted EBITDA", 
    "Property, Plant and Equipment, net", 
    "Other Breakdown"
]

class CustomJSONEncoder(json.JSONEncoder):
    """Custom JSON encoder to handle datetime objects."""
    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.isoformat()  # Convert datetime to ISO 8601 string
        return super().default(obj)

def extract_cell_data(cell):
    """Extracts value, formula, comment, and link from a cell."""
    print("extract_cell_data")
    print(cell.data_type)
    print(cell.comment)
    value = cell.value
    formula = None
    if cell.data_type == "f":  # If cell contains a formula
        formula = cell.value
    comment = cell.comment.text if cell.comment else None

    # Extract hyperlink if it exists
    link = None
    if cell.hyperlink:
        link = cell.hyperlink.target

    return {"value": value, "formula": formula, "comment": comment, "link": link}

def extract_cell_styling(cell, is_source=False, is_formula=False):
    """Extracts styling information from a cell based on custom rules."""
    styling = {
        "text_color": "#000000",  # Default to black
        "text_bold": False,
        "background_color": "#FFFFFF",  # Default to white
        "indents": 0,
        "border": "none"
    }

    # Rule-based text color
    if is_source or not is_formula:  # Source or hardcoded value
        styling["text_color"] = "#0000FF"  # Blue text

    # Bold text if specified in the font
    styling["text_bold"] = cell.font.bold if cell.font else False

    # Background color handling
    if cell.fill and cell.fill.fgColor:
        if cell.fill.fgColor.type == "rgb" and cell.fill.fgColor.rgb:
            rgb_value = cell.fill.fgColor.rgb
            if len(rgb_value) == 8:  # Ensure valid ARGB (Alpha included)
                styling["background_color"] = f"#{rgb_value[2:]}"  # Strip the Alpha channel (leading 2 chars)
            elif len(rgb_value) == 6:  # Handle RGB without Alpha
                styling["background_color"] = f"#{rgb_value}"
            else:
                styling["background_color"] = "#FFFFFF"  # Fallback to white for invalid color
        else:
            styling["background_color"] = "#FFFFFF"  # Default to white if not RGB
    else:
        styling["background_color"] = "#FFFFFF"  # Default to white if no fill color

    # Indentation from alignment
    if cell.alignment:
        styling["indents"] = cell.alignment.indent if cell.alignment.indent else 0

    # Border handling
    if cell.border:
        border_styles = []
        for side_name in ["left", "right", "top", "bottom"]:
            side: Side = getattr(cell.border, side_name)
            if side and side.style:
                border_styles.append(f"{side_name}: {side.style}")
        styling["border"] = ", ".join(border_styles) if border_styles else "none"

    return styling

def is_next_row_empty(sheet, current_row_index):
    """
    Checks if the next row's first column is empty.
    """
    # Handle None or invalid row index
    if current_row_index is None or current_row_index <= 0:
        return False

    # Get the next row index
    next_row_index = current_row_index + 1

    # Check if the next row index is within the valid range
    if next_row_index > sheet.max_row:
        return True  # If there is no next row, consider it empty

    # Check the value in the first column of the next row
    next_row_value = sheet.cell(row=next_row_index, column=1).value

    # Return True if the next row's first column is empty, False otherwise
    return next_row_value is None or str(next_row_value).strip() == ""


def csv_to_nosql(sheet, company, ticker, updated_at, last_updated_with):
    print("Starting Excel to NoSQL transformation...")

    # Extract headers for calendar, fiscal, and fiscal_date
    calendar = [sheet.cell(row=1, column=i).value for i in range(5, sheet.max_column + 1)]
    fiscal = [sheet.cell(row=2, column=i).value for i in range(5, sheet.max_column + 1)]
    fiscal_date = [sheet.cell(row=3, column=i).value for i in range(5, sheet.max_column + 1)]

    print(f"Extracted calendar: {calendar}")
    print(f"Extracted fiscal: {fiscal}")
    print(f"Extracted fiscal_date: {fiscal_date}")

    # Initialize the database object
    database_object = {
        "company": company,
        "ticker": ticker,
        "updated_at": updated_at,
        "last_updated_with": last_updated_with,
        "metrics": []
    }

    # Track the current section, category, and subcategory
    section_name = None
    category_name = None
    subcategory_name = None
    subsubcategory_name = None
    
    # Initialize row numbers with 0 instead of None
    section_row = 0
    category_row = 0
    subcategory_row = 0
    subsubcategory_row = 0
    
    # Initialize styling variables as None
    section_styling = None
    category_styling = None
    subcategory_styling = None
    subsubcategory_styling = None

    # Process rows starting from the 4th row
    for row_index in range(4, sheet.max_row + 1):
        row_data = [sheet.cell(row=row_index, column=i).value for i in range(1, sheet.max_column + 1)]
        row_first_col = row_data[0]
        prior_row_data = [sheet.cell(row=row_index - 1, column=i).value for i in range(1, sheet.max_column + 1)]

        # Check if the row corresponds to a section
        if row_first_col in sections:
            section_name = row_first_col
            section_row = row_index
            category_name = None  # Reset category and subcategory
            subcategory_name = None
            section_styling = extract_cell_styling(sheet.cell(row=row_index, column=1))  # Extract section styling
            print(f"Identified new section: {section_name}")
            continue

        # Check if the row corresponds to a category (empty except for first column)
        if all(value is None for value in row_data[1:]):
            if category_name:
                if subcategory_name:
                    if row_index == subcategory_row + 1:
                        # This row should be a subsubcategory since it's right after a subcategory
                        subsubcategory_name = row_first_col
                        subsubcategory_row = row_index
                        subsubcategory_styling = extract_cell_styling(sheet.cell(row=row_index, column=1))
                        print(f"Identified new subsubcategory: {subsubcategory_name}")
                    else:
                        # This row is a new category since it's not immediately after a subcategory
                        subcategory_name = row_first_col
                        subcategory_row = row_index
                        subcategory_styling = extract_cell_styling(sheet.cell(row=row_index, column=1))
                        # Reset subsubcategory since we have a new category
                        subsubcategory_name = None
                        subsubcategory_row = None
                        subsubcategory_styling = None
                        print(f"Identified new category: {category_name}")
                elif row_index == category_row + 1:
                    # This row should be a subcategory since it's right after a category
                    subcategory_name = row_first_col
                    subcategory_row = row_index
                    subcategory_styling = extract_cell_styling(sheet.cell(row=row_index, column=1))
                    print(f"Identified new subcategory: {subcategory_name}")
                else:
                    # This row is a new category since it's not immediately after a category
                    category_name = row_first_col
                    category_row = row_index
                    category_styling = extract_cell_styling(sheet.cell(row=row_index, column=1))
                    # Reset subcategory and subsubcategory
                    subcategory_name = None
                    subcategory_row = None
                    subcategory_styling = None
                    subsubcategory_name = None
                    subsubcategory_row = None
                    subsubcategory_styling = None
                    print(f"Identified new category: {category_name}")
            else:
                # First category case
                category_name = row_first_col
                category_row = row_index
                category_styling = extract_cell_styling(sheet.cell(row=row_index, column=1))
                print(f"Identified new category: {category_name}")
            continue

        # Process as a metric row
        name_cell = sheet.cell(row=row_index, column=1)
        name_data = extract_cell_data(name_cell)
        name_styling = extract_cell_styling(name_cell, is_formula=False)

        source_cell = sheet.cell(row=row_index, column=3)
        source_data = extract_cell_data(source_cell)
        source_styling = extract_cell_styling(source_cell, is_source=True)  # Source-specific styling

        tag_id = row_data[3]

        # Extract time-series values
        values = []
        for i, col in enumerate(calendar):
            cell = sheet.cell(row=row_index, column=5 + i)  # Time-series columns start from 5th column
            cell_data = extract_cell_data(cell)
            cell_styling = extract_cell_styling(
                cell, 
                is_formula=cell_data["formula"] is not None
            )  # Pass context to styling extraction

            values.append({
                "period": calendar[i],
                "fiscal": fiscal[i],
                "fiscal_date": fiscal_date[i],
                "value": cell_data["value"],
                "formula": cell_data["formula"],
                "comment": cell_data["comment"],
                "link": cell_data["link"],  # Include extracted link
                "styling": cell_styling
            })

        # Build the metric object
        metric = {
            "section": {
                "name": section_name, 
                "order": section_row,
                "styling": section_styling if section_name else None,  # Add section styling
                "empty_row_after": is_next_row_empty(sheet, section_row)
            },
            "category": {
                "name": category_name, 
                "order": category_row,
                "styling": category_styling if category_name else None,  # Add category styling
                "empty_row_after": is_next_row_empty(sheet, category_row)
            },
            "subcategory": {
                "name": subcategory_name, 
                "order": subcategory_row,
                "styling": subcategory_styling if subcategory_name else None,  # Add subcategory styling
                "empty_row_after": is_next_row_empty(sheet, subcategory_row)
            },
            "subsubcategory": {
                "name": subsubcategory_name,
                "order": subsubcategory_row,
                "styling": subsubcategory_styling if subsubcategory_name else None,
                "empty_row_after": is_next_row_empty(sheet, subsubcategory_row)
            },
            "name": {
                "name": name_data["value"],
                "order": row_index,
                "link": name_data["link"],
                "styling": name_styling,  # Apply name styling
                "empty_row_after": is_next_row_empty(sheet, row_index)  # Add empty_row_after parameter
            },
            "values": values,
            "unit": row_data[1],
            "source": {
                "value": source_data["value"],
                "link": source_data["link"],
                "styling": source_styling  # Apply source styling
            },
            "tag_id": tag_id
        }
        database_object["metrics"].append(metric)

    print("Excel to NoSQL transformation complete.")
    return database_object


MONGO_URI = "mongodb+srv://crmchattie:wlfkYHdCsO3MbAz2@daloopa.strbq.mongodb.net/?retryWrites=true&w=majority&appName=Daloopa"
DATABASE_NAME = "daloopa"
COLLECTION_NAME = "companies"

@app.get("/process_file/")
def process_file():
    # Path to the Excel file in your repository
    excel_path = "RDDT Model.xlsx"
    company = "Reddit"
    ticker = "RDDT"
    updated_at = datetime.utcnow()  # Use current UTC time
    last_updated_with = {
        "source": "S-1 Filing",
        "file_link": "https://www.sec.gov/Archives/edgar/data/1713445/000162828024006294/reddits-1q423.htm"
    }

    # Load the Excel file
    try:
        print(f"Attempting to load Excel file from: {excel_path}")
        workbook = load_workbook(excel_path, data_only=False)  # Load with formulas, comments, and styling
        sheet = workbook.active  # Use the first sheet
        print("Excel file loaded successfully.")
    except FileNotFoundError:
        error_message = f"File not found: {excel_path}"
        print(error_message)
        return {"success": False, "error": error_message}
    except Exception as e:
        error_message = f"Error loading Excel file: {str(e)}"
        print(error_message)
        return {"success": False, "error": error_message}

    # Transform the data
    try:
        print("Starting transformation of Excel data...")
        result = csv_to_nosql(sheet, company, ticker, updated_at, last_updated_with)
        print("Transformation completed successfully.")

        # Save the JSON output to a file in the local repository
        output_file = "RDDT_Model_Output.json"
        try:
            with open(output_file, "w", encoding="utf-8") as json_file:
                json.dump(result, json_file, ensure_ascii=False, indent=4, cls=CustomJSONEncoder)
            print(f"JSON data saved to file: {output_file}")
        except Exception as file_error:
            print(f"Error saving JSON to file: {str(file_error)}")

        # Connect to MongoDB
        client = MongoClient(MONGO_URI)
        db = client[DATABASE_NAME]
        collection = db[COLLECTION_NAME]

        # Store the data in MongoDB
        filter_query = {"ticker": ticker, "company": company}  # Define the document to overwrite
        update_data = result  # The data to replace the existing document with
        collection.replace_one(filter_query, update_data, upsert=True)  # Overwrite or insert if not found
        print(f"Data for '{company}' with ticker '{ticker}' replaced in MongoDB collection '{COLLECTION_NAME}'.")

        return {"success": True, "message": "Data stored in MongoDB and saved locally as JSON"}
    except Exception as e:
        error_message = f"Error during transformation: {str(e)}"
        print(error_message)
        return {"success": False, "error": error_message}